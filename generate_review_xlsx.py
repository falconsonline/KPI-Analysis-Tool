#!/usr/bin/env python3
"""Generate KPI Analysis Tool review spreadsheet."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Style Definitions ───────────────────────────────────────────────
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
OK_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
BUG_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
NEW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
EXISTS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
CRIT_FILL = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
HIGH_FILL = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
MED_FILL = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
LOW_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
SECTION_FONT = Font(name='Calibri', bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WRAP = Alignment(wrap_text=True, vertical='top')

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if fill:
            cell.fill = fill

def add_section_row(ws, row, text, cols):
    ws.cell(row=row, column=1, value=text)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

# ══════════════════════════════════════════════════════════════════════
# Sheet 1: Features
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Features"
ws1.sheet_properties.tabColor = "2F5496"

headers = ["ID", "Feature", "Function(s)", "Component", "Status", "Bug ID"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, 1, len(headers))

frontend_features = [
    ("F1", "File Upload — Click", "dataFile change listener", "HTML", "OK", ""),
    ("F2", "File Upload — Drag & Drop", "drop, dragover, dragleave listeners", "HTML", "OK", ""),
    ("F3", "Multi-File Upload", "rawFilesStore accumulation", "HTML", "OK", ""),
    ("F4", "File Status Indicator", "#file-status text update", "HTML", "OK", ""),
    ("F5", "URL Auto-Fetch — Start", "startAutoFetch()", "HTML", "BUG", "H5"),
    ("F6", "URL Auto-Fetch — Stop", "stopAutoFetch()", "HTML", "OK", ""),
    ("F7", "URL Auto-Fetch — Refresh Interval", "fetchInterval + setInterval", "HTML", "OK", ""),
    ("F8", "Log-Pipe Parser", "processFilesFromUI() logPipe", "HTML", "OK", ""),
    ("F9", "Flat CSV Parser", "processFilesFromUI() flatCsv", "HTML", "OK", ""),
    ("F10", "CSV Header Row Detection", "csvHasHeader checkbox", "HTML", "OK", ""),
    ("F11", "CSV Manual Headers", "csvManualHeaders input", "HTML", "OK", ""),
    ("F12", "KPI Pattern Filter — Text", "kpiPattern input", "HTML", "OK", ""),
    ("F13", "KPI Pattern Filter — File", "handlePatternUpload()", "HTML", "OK", ""),
    ("F14", "KPI Pattern with Label", "inputVal.includes(':')", "HTML", "OK", ""),
    ("F15", "Non-printable Char Stripping", "Regex replacement", "HTML", "OK", ""),
    ("F16", "Day-of-Week Prefix Handling", "getUnix() regex strip", "HTML", "OK", ""),
    ("F17", "Timestamp Parsing — Log-Pipe", "getUnix() pipe branch", "HTML", "BUG", "M1"),
    ("F18", "Timestamp Parsing — CSV", "getUnix() comma branch", "HTML", "OK", ""),
    ("F19", "Timestamp Formatting — Log-Pipe", "formatLabel() pipe branch", "HTML", "BUG", "M2"),
    ("F20", "Timestamp Formatting — CSV", "formatLabel() comma branch", "HTML", "OK", ""),
    ("F21", "File Inventory Table", "processFilesFromUI() DOM", "HTML", "OK", ""),
    ("F22", "File Selection — Analyze", "selectFile()", "HTML", "OK", ""),
    ("F23", "Active Row Highlighting", "active-row class", "HTML", "OK", ""),
    ("F24", "Merge All Data", "mergeAllAndSelect()", "HTML", "OK", ""),
    ("F25", "Merged Data Sort by Timestamp", "all.sort()", "HTML", "OK", ""),
    ("F26", "Rule Builder — Add Rule", "addRuleToQueue()", "HTML", "OK", ""),
    ("F27", "Rule Builder — Chart Title", "gHeader input", "HTML", "OK", ""),
    ("F28", "Rule Builder — Mode Direct", "analysisMode = direct", "HTML", "OK", ""),
    ("F29", "Rule Builder — Mode Ratio", "analysisMode = ratio", "HTML", "OK", ""),
    ("F30", "Rule Builder — Multi-Select", "selA multiple select", "HTML", "OK", ""),
    ("F31", "Rule Builder — Warn Threshold", "threshWarn input", "HTML", "OK", ""),
    ("F32", "Rule Builder — Critical Threshold", "threshCritical input", "HTML", "OK", ""),
    ("F33", "Rule Queue Display", "renderQueue()", "HTML", "OK", ""),
    ("F34", "Rule Queue — Remove Rule", "renderQueue splice", "HTML", "OK", ""),
    ("F35", "Rule Queue — Threshold Badge", "hasThresh badge", "HTML", "OK", ""),
    ("F36", "Rule Export to JSON", "saveRulesToFile()", "HTML", "OK", ""),
    ("F37", "Rule Import from JSON", "importRules()", "HTML", "OK", ""),
    ("F38", "Legacy Schema Normalization", "type->mode in importRules", "HTML", "OK", ""),
    ("F39", "Import — Auto-Bind Single File", "importRules single-file", "HTML", "OK", ""),
    ("F40", "Import — CSV Config Restore", "importRules csvHeaderConfig", "HTML", "OK", ""),
    ("F41", "Dashboard Generation", "generateDashboard()", "HTML", "BUG", "H1"),
    ("F42", "Chart.js Line Charts", "Chart constructor", "HTML", "OK", ""),
    ("F43", "Threshold-Colored Points", "pointColors logic", "HTML", "OK", ""),
    ("F44", "Synchronized Zoom/Pan", "syncCharts()", "HTML", "OK", ""),
    ("F45", "Reset View", "resetAllZoom()", "HTML", "OK", ""),
    ("F46", "Dashboard Meta Bar", "active-file-label", "HTML", "OK", ""),
    ("F47", "Ratio Mode Calculation", "tokens[t]/tokens[0]*100", "HTML", "BUG", "L1"),
    ("F48", "Statistics — computeStats", "computeStats()", "HTML", "BUG", "C1"),
    ("F49", "Statistics — Render Table", "renderStats()", "HTML", "OK", ""),
    ("F50", "Statistics — Collapse/Expand", "toggleStatsPanel()", "HTML", "OK", ""),
    ("F51", "Threshold Alert Check", "checkThresholds()", "HTML", "OK", ""),
    ("F52", "Alert Panel — Render", "renderAlerts()", "HTML", "BUG", "H3"),
    ("F53", "Alert Counter Badge", "alertCount textContent", "HTML", "OK", ""),
    ("F54", "Alert Clear", "clearAlerts()", "HTML", "OK", ""),
    ("F55", "WebSocket Client — Connect", "connectWebSocket()", "HTML", "BUG", "H4"),
    ("F56", "WebSocket — Data Messages", "ws.onmessage data", "HTML", "OK", ""),
    ("F57", "WebSocket — Alert Messages", "ws.onmessage alert", "HTML", "OK", ""),
    ("F58", "WebSocket — Live Badge", "live-badge show/hide", "HTML", "OK", ""),
    ("F59", "WebSocket — Auto-Reconnect", "ws.onclose setTimeout", "HTML", "BUG", "H4"),
    ("F60", "WebSocket — Exponential Backoff", "wsDelay doubling", "HTML", "OK", ""),
    ("F61", "WebSocket — URL from Query Param", "?ws= in onload", "HTML", "OK", ""),
    ("F62", "Theme Toggle", "toggleTheme()", "HTML", "OK", ""),
    ("F63", "PDF Export", "exportFullPDF()", "HTML", "OK", ""),
    ("F64", "PDF — Custom Filename", "prompt() dialog", "HTML", "OK", ""),
    ("F65", "PDF — Multi-Page", "y > 250 page break", "HTML", "OK", ""),
    ("F66", "Console — Log Action", "logAction()", "HTML", "BUG", "H2"),
    ("F67", "Console — Toggle Collapse", "toggleConsole()", "HTML", "OK", ""),
    ("F68", "Console — Copy to Clipboard", "copyConsole()", "HTML", "OK", ""),
    ("F69", "Hard Reset", "location.reload()", "HTML", "OK", ""),
    ("F70", "XSS-Safe Filename Display", "DOM textContent", "HTML", "OK", ""),
    ("F71", "XSS-Safe Chart Titles", "textContent in gen", "HTML", "OK", ""),
    ("F72", "Populate Metric Selects", "populateSelects()", "HTML", "OK", ""),
    ("F73", "CSV Toggle Options", "toggleCsvOptions()", "HTML", "OK", ""),
    ("F74", "Duplicate Header Detection", "manualTokens match-skip", "HTML", "OK", ""),
]

server_features = [
    ("S1", "Config Loading", "load_config()", "Server", "BUG", "C4"),
    ("S2", "Config CLI Flag", "sys.argv --config", "Server", "OK", ""),
    ("S3", "Log Parse — Log-Pipe Mode", "parse_log_line() logPipe", "Server", "OK", ""),
    ("S4", "Log Parse — Flat CSV Mode", "parse_log_line() flatCsv", "Server", "BUG", "C7"),
    ("S5", "Pattern Filter (Server)", "pattern check", "Server", "OK", ""),
    ("S6", "Threshold Check (Server)", "check_thresholds()", "Server", "OK", ""),
    ("S7", "Alert — Write to Log", "write_alert_log()", "Server", "BUG", "H7"),
    ("S8", "Alert — Email via SMTP", "send_email_alert()", "Server", "BUG", "C3"),
    ("S9", "File Tail — Local", "tail_local_file()", "Server", "BUG", "C5"),
    ("S10", "File Rotation Detection", "size < pos check", "Server", "BUG", "L2"),
    ("S11", "Remote Sync — rsync", "sync_remote_file()", "Server", "OK", ""),
    ("S12", "Remote Sync — Periodic", "remote_interval timer", "Server", "OK", ""),
    ("S13", "WebSocket Broadcast", "broadcast()", "Server", "BUG", "C2"),
    ("S14", "WebSocket Handler", "ws_handler()", "Server", "BUG", "C6"),
    ("S15", "HTTP Server", "make_http_handler()", "Server", "OK", ""),
    ("S16", "Monitor Loop", "monitor_loop()", "Server", "OK", ""),
    ("S17", "Main Entry", "main(), main_async()", "Server", "BUG", "H8"),
]

r = 2
add_section_row(ws1, r, "FRONTEND FEATURES (KPI-Analysis-Tool.html)", len(headers))
r += 1
for feat in frontend_features:
    for c, v in enumerate(feat, 1):
        ws1.cell(row=r, column=c, value=v)
    fill = BUG_FILL if feat[4] == "BUG" else OK_FILL
    style_row(ws1, r, len(headers), fill=None)
    ws1.cell(row=r, column=5).fill = fill
    r += 1

add_section_row(ws1, r, "SERVER FEATURES (server/server.py)", len(headers))
r += 1
for feat in server_features:
    for c, v in enumerate(feat, 1):
        ws1.cell(row=r, column=c, value=v)
    fill = BUG_FILL if feat[4] == "BUG" else OK_FILL
    style_row(ws1, r, len(headers), fill=None)
    ws1.cell(row=r, column=5).fill = fill
    r += 1

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 10

# ══════════════════════════════════════════════════════════════════════
# Sheet 2: Test Cases
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Test Cases")
ws2.sheet_properties.tabColor = "00B050"

tc_headers = ["TC#", "Group", "Test Case", "Input", "Expected Output", "Status", "Feature ID"]
for c, h in enumerate(tc_headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, len(tc_headers))

test_cases = [
    # Group 1: getUnix
    ("1.1", "getUnix()", "Log-pipe format returns positive integer", "'Jan 15 | 08:00:00'", "> 0", "Exists", "F17"),
    ("1.2", "getUnix()", "Day prefix stripped — same result", "'Mon Jan 15 | 08:00:00' vs 'Jan 15 | 08:00:00'", "Equal", "Exists", "F16"),
    ("1.3", "getUnix()", "CSV format returns positive integer", "'2026-01-15,08:00:00'", "> 0", "Exists", "F18"),
    ("1.4", "getUnix()", "Empty string returns 0", "''", "0", "Exists", "F17"),
    ("1.5", "getUnix()", "Null returns 0", "null", "0", "Exists", "F17"),
    ("1.6", "getUnix()", "Earlier < later timestamp", "08:00 vs 09:00", "earlier < later", "Exists", "F17"),
    ("1.7", "getUnix()", "Uses current year", "Extract year", "Current year", "Exists", "F17"),
    ("1.8", "getUnix()", "Unknown month defaults gracefully", "'Xyz 15 | 08:00:00'", "Uses '01'", "NEW", "F17"),
    ("1.9", "getUnix()", "Missing time part after pipe", "'Jan 15 |'", "Integer (empty time)", "NEW", "F17"),
    ("1.10", "getUnix()", "No pipe or comma format", "'Jan 15 08:00:00'", "0", "NEW", "F17"),
    ("1.11", "getUnix()", "Single element after split", "'Jan| 08:00:00'", "No crash", "NEW", "F17"),
    ("1.12", "getUnix()", "Undefined input", "undefined", "0", "NEW", "F17"),
    ("1.13", "getUnix()", "Different day prefixes stripped", "'Wed Feb 28 | 12:00:00'", "Same as without Wed", "NEW", "F16"),
    ("1.14", "getUnix()", "Milliseconds truncated", "'Jan 15 | 08:00:00.123'", "Same as 08:00:00", "NEW", "F17"),
    ("1.15", "getUnix()", "CSV slashes instead of dashes", "'2026/01/15,08:00:00'", "No crash", "NEW", "F18"),
    # Group 2: formatLabel
    ("2.1", "formatLabel()", "Log-pipe readable label", "'Mon Jan 15 | 14:35:42'", "'Jan 15 14:35'", "Exists", "F19"),
    ("2.2", "formatLabel()", "CSV readable label", "'2026-01-15,14:35:42'", "'2026-01-15 14:35'", "Exists", "F20"),
    ("2.3", "formatLabel()", "Empty string returns empty", "''", "''", "Exists", "F19"),
    ("2.4", "formatLabel()", "Null returns empty", "null", "''", "NEW", "F19"),
    ("2.5", "formatLabel()", "No pipe/comma returns as-is", "'random text'", "'random text'", "NEW", "F19"),
    ("2.6", "formatLabel()", "Pipe with no time", "'Jan 15 |'", "No crash", "NEW", "F19"),
    ("2.7", "formatLabel()", "CSV no time after comma", "'2026-01-15,'", "No crash", "NEW", "F20"),
    ("2.8", "formatLabel()", "Multiple pipes", "'A | B | C'", "First pipe split", "NEW", "F19"),
    # Group 3: Log-Pipe Parser
    ("3.1", "Log-Pipe Parser", "Valid lines parsed correctly", "sample-logpipe.log", "MSU=100", "Exists", "F8"),
    ("3.2", "Log-Pipe Parser", "< 5 pipe segments rejected", "4-segment line", "Not in store", "Exists", "F8"),
    ("3.3", "Log-Pipe Parser", "Pattern filter excludes", "NONEXISTENT", "Empty", "Exists", "F12"),
    ("3.4", "Log-Pipe Parser", "Day prefix doesn't break", "'Mon' prefix", "Parsed ok", "Exists", "F8"),
    ("3.5", "Log-Pipe Parser", "Multiple KV pairs all parsed", "5 KV pairs", "5 tokens", "NEW", "F8"),
    ("3.6", "Log-Pipe Parser", "Odd KV count (trailing key)", "K1,100,K2", "K1=100", "NEW", "F8"),
    ("3.7", "Log-Pipe Parser", "Spaces trimmed from keys", "' K1 , 100 '", "K1=100", "NEW", "F8"),
    ("3.8", "Log-Pipe Parser", "Non-numeric value = 0", "K1,abc", "val=0", "NEW", "F8"),
    ("3.9", "Log-Pipe Parser", "Empty KV section", "Empty 5th segment", "Row filtered", "NEW", "F8"),
    ("3.10", "Log-Pipe Parser", "Non-printable chars stripped", "\\x01\\x02 in line", "Removed", "NEW", "F15"),
    ("3.11", "Log-Pipe Parser", "Pattern with colon label", "'App:PATTERN'", "Filter applied", "NEW", "F14"),
    ("3.12", "Log-Pipe Parser", "Multiple files independently", "2 files", "Both parsed", "NEW", "F3"),
    ("3.13", "Log-Pipe Parser", "Empty file not stored", "Empty string", "Key absent", "NEW", "F8"),
    ("3.14", "Log-Pipe Parser", "Blank lines not stored", "'\\n\\n\\n'", "Key absent", "NEW", "F8"),
    # Group 4: CSV Parser
    ("4.1", "CSV Parser", "Manual headers produce named tokens", "CSV + headers", "Named keys", "Exists", "F9"),
    ("4.2", "CSV Parser", "< 3 columns rejected", "2-column line", "Not in store", "Exists", "F9"),
    ("4.3", "CSV Parser", "Timestamp cols excluded", "Date,Time header", "Not in tokens", "Exists", "F9"),
    ("4.4", "CSV Parser", "Auto-header (csvHasHeader=true)", "Header row", "Names as keys", "NEW", "F10"),
    ("4.5", "CSV Parser", "Manual headers override auto", "Both set", "Manual wins", "NEW", "F11"),
    ("4.6", "CSV Parser", "No headers = Col_N", "No config", "Col_2, Col_3", "NEW", "F9"),
    ("4.7", "CSV Parser", "Duplicate header detection", "Matching data", "Row skipped", "NEW", "F74"),
    ("4.8", "CSV Parser", "Non-numeric values = 0", "abc,def", "0", "NEW", "F9"),
    ("4.9", "CSV Parser", "Trailing comma handled", "Extra comma", "No crash", "NEW", "F9"),
    ("4.10", "CSV Parser", "Windows CRLF line endings", "\\r\\n content", "Parsed ok", "NEW", "F9"),
    # Group 5: Rule Queue
    ("5.1", "Rule Queue", "Add rule grows queue by 1", "Select + add", "length=1", "Exists", "F26"),
    ("5.2", "Rule Queue", "Threshold values stored", "Warn=5, Crit=10", "Populated", "Exists", "F31"),
    ("5.3", "Rule Queue", "DOM entry per rule", "1 rule", "1 div", "Exists", "F33"),
    ("5.4", "Rule Queue", "Remove updates queue+DOM", "Remove 1 of 2", "length=1", "Exists", "F34"),
    ("5.5", "Rule Queue", "No metric = not added", "Empty selection", "Unchanged", "NEW", "F26"),
    ("5.6", "Rule Queue", "Multiple metrics in one rule", "3 selected", "tokens.length=3", "NEW", "F30"),
    ("5.7", "Rule Queue", "Warn-only threshold", "Warn=5 only", "Only warn", "NEW", "F31"),
    ("5.8", "Rule Queue", "Critical-only threshold", "Crit=10 only", "Only critical", "NEW", "F32"),
    ("5.9", "Rule Queue", "No thresholds = empty {}", "Both empty", "{}", "NEW", "F26"),
    ("5.10", "Rule Queue", "Custom title preserved", "'My Chart'", "header='My Chart'", "NEW", "F27"),
    ("5.11", "Rule Queue", "Default title when empty", "''", "header='KPI'", "NEW", "F27"),
    ("5.12", "Rule Queue", "Threshold badge shown", "With thresholds", "Badge in DOM", "NEW", "F35"),
    ("5.13", "Rule Queue", "No badge without thresholds", "Without", "No badge", "NEW", "F35"),
    # Group 6: Rule Import/Export
    ("6.1", "Import/Export", "DBW.json loads 5 rules", "DBW.json", "5 rules", "Exists", "F37"),
    ("6.2", "Import/Export", "Legacy schema normalized", "type:'ratio'", "mode='ratio'", "Exists", "F38"),
    ("6.3", "Import/Export", "GTPP.json loads 3 rules", "GTPP.json", "3 rules", "NEW", "F37"),
    ("6.4", "Import/Export", "CSV config restored", "With csvHeaderConfig", "Populated", "NEW", "F40"),
    ("6.5", "Import/Export", "Missing thresholds = {}", "No thresholds key", "{}", "NEW", "F37"),
    ("6.6", "Import/Export", "Invalid JSON handled", "Malformed", "Alert, no crash", "NEW", "F37"),
    ("6.7", "Import/Export", "Flat array recognized", "[{header:...}]", "Rules loaded", "NEW", "F37"),
    ("6.8", "Import/Export", "Export valid JSON", "saveRulesToFile()", "Blob created", "NEW", "F36"),
    ("6.9", "Import/Export", "Auto-bind single file", "1 file in store", "selectFile called", "NEW", "F39"),
    ("6.10", "Import/Export", "Dashboard triggered", "Data exists", "Dashboard shown", "NEW", "F37"),
    # Group 7: Data Selection & Merge
    ("7.1", "Selection/Merge", "Merge sorted by timestamp", "2 files", "Sorted merged", "Exists", "F24"),
    ("7.2", "Selection/Merge", "Select sets activeLogData", "Click analyze", "Data matches", "NEW", "F22"),
    ("7.3", "Selection/Merge", "Select sets activeFileName", "Click analyze", "Name matches", "NEW", "F22"),
    ("7.4", "Selection/Merge", "Active row highlighted", "Select index 1", "active-row class", "NEW", "F23"),
    ("7.5", "Selection/Merge", "Previous row unhighlighted", "A then B", "Only B active", "NEW", "F23"),
    ("7.6", "Selection/Merge", "Builder panel shown", "Select file", "Visible", "NEW", "F22"),
    ("7.7", "Selection/Merge", "Auto-generate if rules exist", "Rules queued", "Dashboard shown", "NEW", "F41"),
    ("7.8", "Selection/Merge", "Merge auto-generates", "Rules queued", "Dashboard shown", "NEW", "F24"),
    ("7.9", "Selection/Merge", "Merge empty store", "No files", "activeLogData=[]", "NEW", "F24"),
    # Group 8: Dashboard
    ("8.1", "Dashboard", "Canvas per rule", "1 rule", "1 canvas", "Exists", "F42"),
    ("8.2", "Dashboard", "XSS-safe title", "<script> tag", "Text only", "Exists", "F71"),
    ("8.3", "Dashboard", "Multiple rules = multiple wrappers", "3 rules", "3 wrappers", "NEW", "F41"),
    ("8.4", "Dashboard", "Ratio mode correct", "100, 50", "50%", "NEW", "F47"),
    ("8.5", "Dashboard", "Ratio zero denominator = 0", "tokens[0]=0", "0", "NEW", "F47"),
    ("8.6", "Dashboard", "Ratio first token = 100%", "Any", "100%", "NEW", "F47"),
    ("8.7", "Dashboard", "Direct mode raw values", "token=42", "42", "NEW", "F28"),
    ("8.8", "Dashboard", "Dashboard clears on rebuild", "Generate twice", "Old removed", "NEW", "F41"),
    ("8.9", "Dashboard", "chartInstances correct", "2 rules", "length=2", "NEW", "F42"),
    ("8.10", "Dashboard", "Meta bar filename", "test.log", "Label=test.log", "NEW", "F46"),
    ("8.11", "Dashboard", "Empty data = error logged", "No data", "'No data' logged", "NEW", "F41"),
    ("8.12", "Dashboard", "Auto-bind single file", "1 file", "selectFile called", "NEW", "F41"),
    ("8.13", "Dashboard", "Critical coloring (red)", "val>=critical", "#dc3545", "NEW", "F43"),
    ("8.14", "Dashboard", "Warning coloring (yellow)", "warn<=val<crit", "#ffc107", "NEW", "F43"),
    ("8.15", "Dashboard", "No threshold = 0 radius", "No thresholds", "pointRadius=0", "NEW", "F43"),
    # Group 9: Statistics
    ("9.1", "Statistics", "Correct min/max/mean", "[10,20,30]", "10/30/20", "Exists", "F48"),
    ("9.2", "Statistics", "Panel visible after dashboard", "Generate", "Visible", "Exists", "F49"),
    ("9.3", "Statistics", "Empty array = {}", "[]", "{}", "NEW", "F48"),
    ("9.4", "Statistics", "Single point stddev=0", "[42]", "stddev=0", "NEW", "F48"),
    ("9.5", "Statistics", "Negative values", "[-10,-5,0,5]", "min=-10", "NEW", "F48"),
    ("9.6", "Statistics", "Mixed undefined/NaN filtered", "[10,undef,NaN,30]", "[10,30]", "NEW", "F48"),
    ("9.7", "Statistics", "Std deviation correct", "[2,4,4,4,5,5,7,9]", "~2.0", "NEW", "F48"),
    ("9.8", "Statistics", "Table rows = token count", "5 tokens", "5 rows", "NEW", "F49"),
    ("9.9", "Statistics", "Collapse hides body", "Click", "display=none", "NEW", "F50"),
    ("9.10", "Statistics", "Expand shows body", "Click again", "display=''", "NEW", "F50"),
    ("9.11", "Statistics", "Multiple tokens in table", "3 metrics", "All 3 shown", "NEW", "F49"),
    ("9.12", "Statistics", "Large dataset no crash", "150K values", "No crash", "NEW", "F48"),
    # Group 10: Alerts
    ("10.1", "Alerts", "Value >= critical", "val=15, crit=10", "critical", "Exists", "F51"),
    ("10.2", "Alerts", "warn <= val < critical", "val=7, w=5, c=10", "warn", "Exists", "F51"),
    ("10.3", "Alerts", "Value < warn = none", "val=3, w=5", "0 breaches", "Exists", "F51"),
    ("10.4", "Alerts", "Panel visible on breach", "Breach", "Visible", "Exists", "F52"),
    ("10.5", "Alerts", "Exactly at warn", "val=5, warn=5", "Breach", "NEW", "F51"),
    ("10.6", "Alerts", "Exactly at critical", "val=10, crit=10", "Critical", "NEW", "F51"),
    ("10.7", "Alerts", "Mixed breach/no-breach", "2 metrics", "1 breach", "NEW", "F51"),
    ("10.8", "Alerts", "No thresholds = none", "thresholds={}", "0 breaches", "NEW", "F51"),
    ("10.9", "Alerts", "Metric not in data", "X not in data", "0 breaches", "NEW", "F51"),
    ("10.10", "Alerts", "Counter correct", "3 breaches", "alertCount=3", "NEW", "F53"),
    ("10.11", "Alerts", "Clear resets all", "clearAlerts()", "Empty + hidden", "NEW", "F54"),
    ("10.12", "Alerts", "Formatted timestamp", "Breach data", "formatLabel output", "NEW", "F52"),
    ("10.13", "Alerts", "Critical CSS class", "Critical", ".critical", "NEW", "F52"),
    ("10.14", "Alerts", "Warning CSS class", "Warn", ".warn", "NEW", "F52"),
    ("10.15", "Alerts", "Multiple rules thresholds", "2 rules", "All checked", "NEW", "F51"),
    # Group 11: UI
    ("11.1", "UI", "Theme toggle cycles", "Twice", "light->dark->light", "Exists", "F62"),
    ("11.2", "UI", "logAction appends", "Message", "Entry in DOM", "Exists", "F66"),
    ("11.3", "UI", "CSV options for flatCsv", "flatCsv", "Visible", "Exists", "F73"),
    ("11.4", "UI", "CSV options hidden logPipe", "logPipe", "Hidden", "Exists", "F73"),
    ("11.5", "UI", "Drop zone dragover", "Event", "drag-over class", "Exists", "F2"),
    ("11.6", "UI", "Console toggle", "Click header", "Toggles class", "NEW", "F67"),
    ("11.7", "UI", "Console arrow direction", "Toggle", "Arrow changes", "NEW", "F67"),
    ("11.8", "UI", "Dark theme attribute", "Toggle dark", "data-theme=dark", "NEW", "F62"),
    ("11.9", "UI", "Log has timestamp", "logAction", "[HH:MM:SS]", "NEW", "F66"),
    ("11.10", "UI", "Dragleave removes class", "Event", "drag-over gone", "NEW", "F2"),
    ("11.11", "UI", "Multiple logs accumulate", "5 messages", "5+ entries", "NEW", "F66"),
    ("11.12", "UI", "Populate selects", "Set tokens", "Options match", "NEW", "F72"),
    # Group 12: Auto-Fetch
    ("12.1", "Auto-Fetch", "Empty URL warning", "No URL", "'No URL' logged", "NEW", "F5"),
    ("12.2", "Auto-Fetch", "Start hides/shows buttons", "Start", "Start hidden", "NEW", "F5"),
    ("12.3", "Auto-Fetch", "Stop shows/hides buttons", "Stop", "Start shown", "NEW", "F6"),
    ("12.4", "Auto-Fetch", "Successful fetch stores data", "Mock URL", "Store updated", "NEW", "F5"),
    ("12.5", "Auto-Fetch", "Failed fetch logs error", "404", "Error logged", "NEW", "F5"),
    ("12.6", "Auto-Fetch", "Custom interval", "10s", "10000ms", "NEW", "F7"),
    ("12.7", "Auto-Fetch", "Default interval 30s", "No change", "30s", "NEW", "F7"),
    ("12.8", "Auto-Fetch", "Stop clears timer", "Start+stop", "null", "NEW", "F6"),
    # Group 13: WebSocket
    ("13.1", "WebSocket", "Query param connects", "?ws=url", "Connected", "NEW", "F61"),
    ("13.2", "WebSocket", "Data message appends", "{type:data}", "Data grows", "NEW", "F56"),
    ("13.3", "WebSocket", "Alert message renders", "{type:alert}", "Alert in DOM", "NEW", "F57"),
    ("13.4", "WebSocket", "Live badge on connect", "onopen", "Badge visible", "NEW", "F58"),
    ("13.5", "WebSocket", "Badge hidden disconnect", "onclose", "Badge hidden", "NEW", "F58"),
    ("13.6", "WebSocket", "Backoff doubles", "Disconnect", "wsDelay=2000", "NEW", "F60"),
    ("13.7", "WebSocket", "Backoff capped 30s", "Many disconnects", "Max 30000", "NEW", "F60"),
    ("13.8", "WebSocket", "Delay resets on connect", "Reconnect ok", "wsDelay=1000", "NEW", "F60"),
    ("13.9", "WebSocket", "Non-JSON handled", "Bad message", "Logged, no crash", "NEW", "F56"),
    ("13.10", "WebSocket", "Metrics populated", "First data", "Tokens set", "NEW", "F56"),
    # Group 14: PDF Export
    ("14.1", "PDF Export", "No charts warning", "0 charts", "'No charts' logged", "NEW", "F63"),
    ("14.2", "PDF Export", "Cancel prompt", "null", "'cancelled' logged", "NEW", "F64"),
    ("14.3", "PDF Export", "Empty filename default", "''", "Default name", "NEW", "F64"),
    ("14.4", "PDF Export", ".pdf auto-appended", "'report'", "'report.pdf'", "NEW", "F64"),
    ("14.5", "PDF Export", "Existing .pdf not doubled", "'report.pdf'", "'report.pdf'", "NEW", "F64"),
    # Group 15: File Upload Events
    ("15.1", "File Upload", "Upload stores content", "1 file", "Key present", "NEW", "F1"),
    ("15.2", "File Upload", "Multiple files", "3 files", "3 keys", "NEW", "F3"),
    ("15.3", "File Upload", "Status shows count", "2 files", "'2 files ready.'", "NEW", "F4"),
    ("15.4", "File Upload", "Triggers processFilesFromUI", "File added", "Parsed", "NEW", "F1"),
    ("15.5", "File Upload", "Triggers merge", "File added", "Merged", "NEW", "F1"),
    # Group 16: Pattern File
    ("16.1", "Pattern File", "Loads into input", "'SYSTEM_KPI'", "Input set", "NEW", "F13"),
    ("16.2", "Pattern File", "Colon label extracted", "'App:PATTERN'", "Label logged", "NEW", "F14"),
    ("16.3", "Pattern File", "Empty file no change", "Empty", "Unchanged", "NEW", "F13"),
    ("16.4", "Pattern File", "Triggers reprocessing", "Files loaded", "Reprocessed", "NEW", "F13"),
    # Group 17: Chart Sync
    ("17.1", "Chart Sync", "Sync propagates zoom", "Zoom chart 0", "Chart 1 matches", "NEW", "F44"),
    ("17.2", "Chart Sync", "isInternalSync prevents loops", "Internal", "No recursion", "NEW", "F44"),
    ("17.3", "Chart Sync", "Reset all zoom", "Multiple charts", "All reset", "NEW", "F45"),
    # Group 18: Edge Cases
    ("18.1", "Edge Cases", "Empty rawFilesStore", "No files", "No crash", "NEW", "F8"),
    ("18.2", "Edge Cases", "Empty ruleQueue", "No rules", "No charts", "NEW", "F41"),
    ("18.3", "Edge Cases", "All same values", "[5,5,5]", "stddev=0", "NEW", "F48"),
    ("18.4", "Edge Cases", "Very large values", "[1e15,2e15]", "Correct", "NEW", "F48"),
    ("18.5", "Edge Cases", "Special chars in token", "'SM TimedOut'", "Handled", "NEW", "F8"),
    ("18.6", "Edge Cases", "XSS filename", "'<img>.log'", "Text only", "NEW", "F70"),
    ("18.7", "Edge Cases", "Very long line (10KB)", "Large line", "No crash", "NEW", "F8"),
    ("18.8", "Edge Cases", "Mixed valid/invalid lines", "Mixed", "Valid parsed", "NEW", "F8"),
]

r = 2
for tc in test_cases:
    for c, v in enumerate(tc, 1):
        ws2.cell(row=r, column=c, value=v)
    fill = EXISTS_FILL if tc[5] == "Exists" else NEW_FILL
    style_row(ws2, r, len(tc_headers))
    ws2.cell(row=r, column=6).fill = fill
    r += 1

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 25
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 10

# ══════════════════════════════════════════════════════════════════════
# Sheet 3: Bugs
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Bugs")
ws3.sheet_properties.tabColor = "FF0000"

bug_headers = ["Bug ID", "Severity", "Component", "Description", "Affected Feature", "Fix Required", "Status"]
for c, h in enumerate(bug_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(bug_headers))

bugs = [
    ("C1", "CRITICAL", "HTML", "computeStats() uses Math.min(...vals) — stack overflow >100K elements", "F48", "Replace with reduce pattern", "Open"),
    ("C2", "CRITICAL", "Server", "broadcast() iterates clients set while ws_handler() modifies it — RuntimeError", "S13", "Iterate over set(clients) snapshot", "Open"),
    ("C3", "CRITICAL", "Server", "Email flooding — every log line above threshold triggers email", "S8", "AlertDeduplicator with 300s cooldown", "Open"),
    ("C4", "CRITICAL", "Server", "load_config() shallow merge — partial email config KeyError crash", "S1", "Recursive deep_merge()", "Open"),
    ("C5", "CRITICAL", "Server", "File positions start at 0 — entire large log loaded OOM", "S9", "start_from_end config, seek to EOF", "Open"),
    ("C6", "CRITICAL", "Server", "ws_server_handler path param removed in websockets >= 10.0", "S14", "Remove path parameter", "Open"),
    ("C7", "CRITICAL", "Server", "CSV mode Col_N names mismatch with manualHeaders in rules", "S4", "Map to header names from config", "Open"),
    ("H1", "HIGH", "HTML", "generateDashboard() and selectFile() mutual recursion", "F41", "isGeneratingDashboard guard flag", "Open"),
    ("H2", "HIGH", "HTML", "logAction() unbounded DOM entries — memory leak", "F66", "Cap at 500 entries", "Open"),
    ("H3", "HIGH", "HTML", "renderAlerts() unlimited DOM elements, no batching", "F52", "DocumentFragment, cap at 1000/200", "Open"),
    ("H4", "HIGH", "HTML", "connectWebSocket() reconnects forever, growing closure chain", "F55,F59", "Max 10 attempts, close old ws", "Open"),
    ("H5", "HIGH", "HTML", "startAutoFetch() double-click creates overlapping timers", "F5", "isFetchRunning boolean guard", "Open"),
    ("H6", "HIGH", "HTML", "File input onchange won't fire for same file twice", "F1", "Reset input.value = ''", "Open"),
    ("H7", "HIGH", "Server", "alerts.log append-only, no rotation — fills disk", "S7", "RotatingFileHandler (10MB, 5 backups)", "Open"),
    ("H8", "HIGH", "Server", "No graceful shutdown — wait_closed() blocks forever", "S17", "Signal handler + try-finally", "Open"),
    ("M1", "MEDIUM", "HTML", "getUnix() missing boundary checks — undefined dateParts crash", "F17", "Length/month validation", "Open"),
    ("M2", "MEDIUM", "HTML", "formatLabel() crash when split < 2 parts", "F19", "Length guards", "Open"),
    ("M3", "MEDIUM", "HTML", "CSV parser naive split breaks quoted fields", "F9", "Document as limitation", "Open"),
    ("M4", "MEDIUM", "HTML", "Drop event may trigger file dialog (click interference)", "F2", "stopPropagation + debounce", "Open"),
    ("M5", "MEDIUM", "HTML", "processFilesFromUI() silent on empty rawFilesStore", "F8", "Early return + warning", "Open"),
    ("M6", "MEDIUM", "Server", "errors='replace' silently corrupts metric names", "S9", "UTF-8 strict, latin-1 fallback", "Open"),
    ("L1", "LOW", "HTML", "Ratio mode no zero-denominator guard", "F47", "Explicit check for 0", "Open"),
    ("L2", "LOW", "Server", "File rotation check unreachable (< after <=)", "S10", "Change <= to ==", "Open"),
]

r = 2
for bug in bugs:
    for c, v in enumerate(bug, 1):
        ws3.cell(row=r, column=c, value=v)
    style_row(ws3, r, len(bug_headers))
    sev = bug[1]
    if sev == "CRITICAL":
        ws3.cell(row=r, column=2).fill = CRIT_FILL
        ws3.cell(row=r, column=2).font = Font(bold=True, color='FFFFFF')
    elif sev == "HIGH":
        ws3.cell(row=r, column=2).fill = HIGH_FILL
        ws3.cell(row=r, column=2).font = Font(bold=True)
    elif sev == "MEDIUM":
        ws3.cell(row=r, column=2).fill = MED_FILL
    else:
        ws3.cell(row=r, column=2).fill = LOW_FILL
    r += 1

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 55
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 40
ws3.column_dimensions['G'].width = 10

# ══════════════════════════════════════════════════════════════════════
# Sheet 4: Security
# ══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Security")
ws4.sheet_properties.tabColor = "7030A0"

sec_headers = ["ID", "Severity", "Component", "Vulnerability", "Fix Required", "Status"]
for c, h in enumerate(sec_headers, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, len(sec_headers))

security = [
    ("SEC1", "CRITICAL", "Server", "Command injection via rsync config (user/host/path fields)", "Validate with strict regex patterns", "Open"),
    ("SEC2", "CRITICAL", "HTML", "WS URL injection via ?ws= query parameter — arbitrary WebSocket", "Validate protocol whitelist (ws:/wss:)", "Open"),
    ("SEC3", "HIGH", "HTML", "No Subresource Integrity on CDN scripts — CDN compromise = takeover", "Add integrity + crossorigin attributes", "Open"),
    ("SEC4", "HIGH", "HTML", "Fetch URL not validated — SSRF to internal services", "Whitelist http/https, block private IPs", "Open"),
    ("SEC5", "HIGH", "Server", "SMTP injection via recipients list — header injection", "Validate email format with regex", "Open"),
    ("SEC6", "HIGH", "Server", "Plaintext SMTP password in config.json", "Use SMTP_PASSWORD environment variable", "Open"),
    ("SEC7", "HIGH", "Server", "No WebSocket origin validation — cross-origin data theft", "Check Origin header against allowed list", "Open"),
    ("SEC8", "HIGH", "Server", "HTTP error leaks filesystem path in error message", "Return generic 'Not found' message", "Open"),
    ("SEC9", "HIGH", "HTML", "No file upload size/type validation — DoS vector", "50MB limit, extension whitelist (.txt,.log,.csv,.json)", "Open"),
    ("SEC10", "MEDIUM", "Server", "World-readable /tmp/kpi-remote temp directory", "tempfile.mkdtemp with 0o700 permissions", "Open"),
    ("SEC11", "MEDIUM", "Server", "Unbounded broadcast message size — DoS", "1MB max message size limit", "Open"),
    ("SEC12", "MEDIUM", "HTML", "Missing Content-Security-Policy meta tag", "Add CSP restricting script sources", "Open"),
]

r = 2
for sec in security:
    for c, v in enumerate(sec, 1):
        ws4.cell(row=r, column=c, value=v)
    style_row(ws4, r, len(sec_headers))
    sev = sec[1]
    if sev == "CRITICAL":
        ws4.cell(row=r, column=2).fill = CRIT_FILL
        ws4.cell(row=r, column=2).font = Font(bold=True, color='FFFFFF')
    elif sev == "HIGH":
        ws4.cell(row=r, column=2).fill = HIGH_FILL
        ws4.cell(row=r, column=2).font = Font(bold=True)
    elif sev == "MEDIUM":
        ws4.cell(row=r, column=2).fill = MED_FILL
    r += 1

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 10
ws4.column_dimensions['D'].width = 55
ws4.column_dimensions['E'].width = 45
ws4.column_dimensions['F'].width = 10

# ══════════════════════════════════════════════════════════════════════
# Sheet 5: Summary
# ══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary")
ws5.sheet_properties.tabColor = "00B0F0"

sum_headers = ["Category", "Count"]
for c, h in enumerate(sum_headers, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, 1, len(sum_headers))

summary_data = [
    ("Frontend Features", 74),
    ("Server Features", 17),
    ("Total Features", 91),
    ("", ""),
    ("Existing Test Cases", 27),
    ("New Test Cases Required", 110),
    ("Total Test Cases", 137),
    ("", ""),
    ("Critical Bugs", 7),
    ("High Bugs", 8),
    ("Medium Bugs", 6),
    ("Low/Logic Bugs", 2),
    ("Total Bugs", 23),
    ("", ""),
    ("Critical Security Vulnerabilities", 2),
    ("High Security Vulnerabilities", 6),
    ("Medium Security Vulnerabilities", 3),
    ("Total Security Vulnerabilities", 12),
]

r = 2
for item in summary_data:
    ws5.cell(row=r, column=1, value=item[0])
    ws5.cell(row=r, column=2, value=item[1])
    style_row(ws5, r, 2)
    if item[0].startswith("Total"):
        ws5.cell(row=r, column=1).font = Font(bold=True)
        ws5.cell(row=r, column=2).font = Font(bold=True)
    r += 1

ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 10

# ── Save ──────────────────────────────────────────────────────────────
import os
output_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(output_dir, "KPI-Analysis-Tool-Review.xlsx")
wb.save(output_path)
print(f"Excel file saved: {output_path}")
